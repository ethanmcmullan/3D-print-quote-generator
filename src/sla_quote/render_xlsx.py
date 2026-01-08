from __future__ import annotations

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font

from .engine import QuoteResult
from .model import QuoteInput

def write_xlsx(outpath: str | Path, q: QuoteInput, r: QuoteResult) -> Path:
    outpath = Path(outpath)
    wb = Workbook()
    ws = wb.active
    ws.title = "Job Log"

    bold = Font(bold=True)
    def row(k, v, rnum):
        ws[f"A{rnum}"] = k
        ws[f"B{rnum}"] = v
        ws[f"A{rnum}"].font = bold

    row("Quote ID", r.quote_id, 1)
    row("Customer", q.customer.name, 2)
    row("Part", q.part.name, 3)
    row("Part Number", q.part.part_number, 4)
    row("Revision", q.part.revision, 5)
    row("Qty", r.qty, 6)

    row("Printer", q.process.printer, 8)
    row("Resin", q.process.resin, 9)
    row("Part Volume (ml)", q.process.part_volume_ml, 10)
    row("Print Hours", q.process.print_hours, 11)

    row("Sell Price (pre-tax)", r.sell_price, 13)
    ws["B13"].number_format = '"$"#,##0.00'
    row("Price / Part", r.price_per_part, 14)
    ws["B14"].number_format = '"$"#,##0.00'
    row("Volume Discount", r.unit_discount_pct, 15)
    ws["B15"].number_format = "0.00%"

    ws["A17"] = "Line Item"; ws["B17"] = "Cost"
    ws["A17"].font = bold; ws["B17"].font = bold

    rr = 18
    for li in r.line_items:
        ws[f"A{rr}"] = li.name
        ws[f"B{rr}"] = li.cost
        ws[f"B{rr}"].number_format = '"$"#,##0.00'
        rr += 1

    wb.save(outpath)
    return outpath
